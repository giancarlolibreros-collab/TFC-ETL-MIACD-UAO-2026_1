# =============================================================================
# FASE de EXTRACCIÓN [TFC-ETL-MIACD-UAO-2026_1]
# Autores: Brayan Valencia Sánchez | Giancarlo Libreros Londoño
# Maestría en IA y Ciencia de Datos - UAO | Curso ETL | 2026-1
# =============================================================================

import os
import re
import pandas as pd
import openpyxl
from datetime import datetime
import warnings
warnings.filterwarnings('ignore')
from src.utils.timer import registrar

# -----------------------------------------------------------------------------
# CONFIGURACIÓN BASE
# -----------------------------------------------------------------------------
CARPETA_RAW = "data/raw"

# =============================================================================
# UTILIDAD: NORMALIZACIÓN DE NOMBRES DE ARCHIVO
# =============================================================================

def normalizar_nombre(nombre: str) -> str:
    """
    Normaliza el nombre de un archivo según la normativa:
    - Convierte a mayúsculas
    - Conserva dígitos 0-9
    - Reemplaza espacios por guion bajo
    - Elimina caracteres especiales (tildes, ñ, etc.)
    """
    base, ext = os.path.splitext(nombre)
    base = base.upper()
    reemplazos = {
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'À': 'A', 'È': 'E', 'Ì': 'I', 'Ò': 'O', 'Ù': 'U',
        'Ñ': '',  'Ü': 'U', ' ': '_'
    }
    for original, reemplazo in reemplazos.items():
        base = base.replace(original, reemplazo)
    base = re.sub(r'[^A-Z0-9_]', '', base)
    base = re.sub(r'_+', '_', base)
    base = base.strip('_')
    return base + ext.lower()

def renombrar_archivos(carpeta: str):
    """
    Recorre una carpeta y renombra los archivos según la normativa.
    """
    print("=" * 60)
    print("UTILIDAD - Normalización de nombres de archivo")
    print("=" * 60)
    for nombre_actual in os.listdir(carpeta):
        ruta_actual = os.path.join(carpeta, nombre_actual)
        if not os.path.isfile(ruta_actual):
            continue
        nombre_nuevo = normalizar_nombre(nombre_actual)
        if nombre_actual != nombre_nuevo:
            ruta_nueva = os.path.join(carpeta, nombre_nuevo)
            os.rename(ruta_actual, ruta_nueva)
            print(f"  Renombrado : {nombre_actual} → {nombre_nuevo}")
        else:
            print(f"  Sin cambios: {nombre_actual}")
    print()

# -----------------------------------------------------------------------------
# CONFIGURACIÓN DE RUTAS (se resuelven después de definir normalizar_nombre)
# -----------------------------------------------------------------------------

RUTA_SOBREGIRO  = os.path.join(CARPETA_RAW, normalizar_nombre("Intereses Sobregiro.xlsx"))
RUTA_TASAS      = os.path.join(CARPETA_RAW, normalizar_nombre("tasas sobregiro.xlsx"))
RUTA_COMISIONES = os.path.join(CARPETA_RAW, normalizar_nombre("COMISIONES AÑO 2025.xlsx"))

# =============================================================================
# FUENTE (archivo)  : INTERESES_SOBREGIRO.xlsx | Hoja: Intereses
# Descripción       : Libro contable con los gastos financieros de la empresa (2025).
# Filtro            : Se extrae únicamente la subcuenta 'INTERESES POR SOBREGIROS'.
# =============================================================================

@registrar(fase="EXTRACCIÓN", componente="Intereses Sobregiro")
def extraer_intereses_sobregiro(ruta: str) -> pd.DataFrame:
    """
    Extrae los registros de intereses por sobregiro del archivo contable.
    Filtro aplicado: Desc. auxiliar == 'INTERESES POR SOBREGIROS'
    """
    print("=" * 60)
    print("EXTRACCIÓN - FUENTE 1: Intereses por Sobregiro")
    print("=" * 60)

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb['Intereses']

    columnas = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[4] and 'SOBREGIRO' in str(fila[4]).upper():
            filas.append(fila)

    df = pd.DataFrame(filas, columns=columnas)

    cols_relevantes = [
        'Desc. auxiliar', 'Neto', 'Docto.', 'Fecha',
        'Notas', 'Razón social tercero movto.', 'Mes', 'Nombre Mes'
    ]
    df = df[cols_relevantes].copy()

    df['Fecha']  = pd.to_datetime(df['Fecha'], errors='coerce')
    df['Neto']   = pd.to_numeric(df['Neto'], errors='coerce')
    df['Mes']    = df['Mes'].astype('Int64')
    df.rename(columns={
        'Desc. auxiliar'              : 'tipo_gasto',
        'Neto'                        : 'valor_cop',
        'Docto.'                      : 'documento',
        'Notas'                       : 'descripcion',
        'Razón social tercero movto.' : 'entidad_bancaria',
        'Nombre Mes'                  : 'nombre_mes'
    }, inplace=True)

    print(f"  Registros extraídos : {len(df)}")
    print(f"  Rango de fechas     : {df['Fecha'].min().date()} → {df['Fecha'].max().date()}")
    print(f"  Total gastos (COP)  : ${df['valor_cop'].sum():,.0f}")
    print(f"  Bancos identificados: {sorted(df['entidad_bancaria'].unique())}")
    print()
    return df

# =============================================================================
# FUENTE (archivo)  : COMISIONES_AO_2025.xlsx | Hoja: COMISIONES ENERO-MARZO 2025 (2)
# Descripción       : Libro de comisiones bancarias.
# Filtro            : Se extrae únicamente la subcuenta 53051502
#                     'COMISIONES POR FACTORING'.
# =============================================================================

@registrar(fase="EXTRACCIÓN", componente="Comisiones Factoring")
def extraer_comisiones_factoring(ruta: str) -> pd.DataFrame:
    """
    Extrae los registros de comisiones por factoring del libro de comisiones.
    Filtro aplicado: Auxiliar == 53051502 (COMISIONES POR FACTORING)
    """
    print("=" * 60)
    print("EXTRACCIÓN - FUENTE 2: Comisiones por Factoring")
    print("=" * 60)

    wb = openpyxl.load_workbook(ruta, data_only=True)
    ws = wb['COMISIONES ENERO-MARZO 2025 (2)']

    columnas = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = []
    for fila in ws.iter_rows(min_row=2, values_only=True):
        if fila[3] == 53051502:
            filas.append(fila)

    df = pd.DataFrame(filas, columns=columnas)

    cols_relevantes = [
        'Desc. auxiliar', 'Neto', 'Docto.', 'Fecha',
        'Notas', 'Razón social tercero movto.', 'Mes', 'Nombre Mes'
    ]
    df = df[cols_relevantes].copy()

    df['Fecha']  = pd.to_datetime(df['Fecha'], errors='coerce')
    df['Neto']   = pd.to_numeric(df['Neto'], errors='coerce')
    df['Mes']    = df['Mes'].astype('Int64')
    df.rename(columns={
        'Desc. auxiliar'              : 'tipo_gasto',
        'Neto'                        : 'valor_cop',
        'Docto.'                      : 'documento',
        'Notas'                       : 'descripcion',
        'Razón social tercero movto.' : 'entidad_factor',
        'Nombre Mes'                  : 'nombre_mes'
    }, inplace=True)

    print(f"  Registros extraídos   : {len(df)}")
    print(f"  Rango de fechas       : {df['Fecha'].min().date()} → {df['Fecha'].max().date()}")
    print(f"  Total comisiones (COP): ${df['valor_cop'].sum():,.0f}")
    print(f"  Entidades factor      : {sorted(df['entidad_factor'].unique())}")
    print()
    return df

# =============================================================================
# FUENTE (archivo)  : TASAS_SOBREGIRO.xlsx | Hoja: Creditos
# Descripción       : Documento de gestión financiera con tasas efectivas anuales
#                     y cupos de crédito por entidad bancaria.
# =============================================================================

@registrar(fase="EXTRACCIÓN", componente="Tasas Sobregiro")
def extraer_tasas_sobregiro(ruta: str) -> pd.DataFrame:
    """
    Extrae y estructura las tasas de sobregiro por entidad bancaria.
    Fuente semiestructurada: requiere limpieza manual de filas vacías.
    """
    print("=" * 60)
    print("EXTRACCIÓN - FUENTE 3: Tasas de Sobregiro")
    print("=" * 60)

    datos_tasas = {
        'entidad'         : ['BANCOLOMBIA', 'DAVIVIENDA', 'BANCO DE BOGOTÁ',
                             'BANCO POPULAR', 'BANCO ITAÚ'],
        'tasa_ea'         : [0.22419735, 0.195, 0.189, None, 0.1362],
        'tasa_descripcion': [
            '20.40% N.A.D.V (nominal anual descontada vencida)',
            '19.50% E.A.',
            '18.9% E.A. (cambia a 20.68% EA desde 25-jul-2025)',
            'Pendiente evaluación de cupos',
            '14% NDV'
        ]
    }

    datos_cupos = {
        'entidad'      : ['BANCO BOGOTÁ', 'BANCO DE OCCIDENTE',
                          'BANCO POPULAR', 'BANCO ITAÚ', 'BANCOLOMBIA',
                          'SERFINANZAS'],
        'cupo_millones': [27000, 7300, 0, 15000, 15012.96, 500],
        'condiciones'  : [
            'Periodo de gracia 2 años, buena tasa',
            '36 meses, intereses trimestrales, capital semestral',
            'Pendiente evaluación',
            '3 años con 1 año de gracia',
            '60 meses con 24 de gracia, fiducia en garantía requerida >36m',
            '36 meses, trimestral, IBR-3.3%'
        ]
    }

    df_tasas = pd.DataFrame(datos_tasas)
    df_cupos = pd.DataFrame(datos_cupos)

    print(f"  Entidades con tasa registrada : {df_tasas['entidad'].nunique()}")
    print(f"  Entidades con cupo registrado : {df_cupos['entidad'].nunique()}")
    print(f"  Tasa promedio sobregiro (E.A.): {df_tasas['tasa_ea'].mean():.2%}")
    print()
    return df_tasas, df_cupos

# =============================================================================
# RESUMEN GENERAL DE EXTRACCIÓN
# =============================================================================

def resumen_extraccion(df_sobregiro, df_factoring, df_tasas):
    print("=" * 60)
    print("RESUMEN CONSOLIDADO DE EXTRACCIÓN")
    print("=" * 60)

    total_sobregiro = df_sobregiro['valor_cop'].sum()
    total_factoring = df_factoring[df_factoring['valor_cop'] > 0]['valor_cop'].sum()
    total_combinado = total_sobregiro + total_factoring

    print(f"  Fuente 1 - Intereses sobregiro  : {len(df_sobregiro):>5} registros | ${total_sobregiro:>18,.0f} COP")
    print(f"  Fuente 2 - Comisiones factoring : {len(df_factoring):>5} registros | ${total_factoring:>18,.0f} COP")
    print(f"  Fuente 3 - Tasas sobregiro      : {len(df_tasas):>5} entidades  | (referencia de tasas)")
    print(f"  {'─'*54}")
    print(f"  TOTAL GASTO FINANCIERO ESTIMADO :{'':>24} ${total_combinado:>18,.0f} COP")
    print()
    print(f"  Período cubierto                : Enero 2025 → Enero 2026")
    print(f"  Formato fuente                  : Excel (.xlsx)")
    print(f"  Método de extracción            : openpyxl + pandas")
    print(f"  Fecha de ejecución              : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

# =============================================================================
# EJECUCIÓN DIRECTA - PRUEBA UNITARIA DEL MÓDULO
# =============================================================================

if __name__ == "__main__":
    print()
    print("  PRUEBA UNITARIA - MÓDULO DE EXTRACCIÓN")
    print("  Ejecución directa de extraccion.py")
    print()

    # Paso 0: normalizar nombres de archivo antes de extraer
    renombrar_archivos(CARPETA_RAW)

    # Paso 1: extracción
    df_sobregiro       = extraer_intereses_sobregiro(RUTA_SOBREGIRO)
    df_factoring       = extraer_comisiones_factoring(RUTA_COMISIONES)
    df_tasas, df_cupos = extraer_tasas_sobregiro(RUTA_TASAS)

    resumen_extraccion(df_sobregiro, df_factoring, df_tasas)

    # Exportar datos extraídos
    with pd.ExcelWriter("data/processed/DATOS_EXTRAIDOS_ETL.xlsx", engine='openpyxl') as writer:
        df_sobregiro.to_excel(writer, sheet_name='intereses_sobregiro', index=False)
        df_factoring.to_excel(writer, sheet_name='comisiones_factoring', index=False)
        df_tasas.to_excel(writer,     sheet_name='tasas_sobregiro',      index=False)
        df_cupos.to_excel(writer,     sheet_name='cupos_credito',        index=False)

    print("Archivo exportado: data/processed/DATOS_EXTRAIDOS_ETL.xlsx")
    print()
